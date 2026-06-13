from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.schemas import AssetKind, WaveSpeedCatalogField, WaveSpeedCatalogModel

REQUIRED_SHEETS = {
    "Models_Full",
    "API_Schemas",
    "Schema_Fields",
    "Capability_Summary",
    "Cheapest_By_Capability",
}

TEXTAREA_NAMES = {
    "prompt",
    "negative_prompt",
    "style_prompt",
    "lyrics",
    "text",
    "description",
    "voice_description",
    "instructions",
    "query",
    "caption",
}

IMAGE_FIELDS = {
    "image",
    "source_image",
    "target_image",
    "last_image",
    "first_frame",
    "reference_image",
    "mask",
    "mask_image",
    "body_image",
    "front_image_url",
    "back_image_url",
    "left_image_url",
    "right_image_url",
}
IMAGE_LIST_FIELDS = {
    "images",
    "source_images",
    "target_images",
    "reference_images",
    "image_urls",
    "reference_urls",
    "refer_images",
    "mask_images",
    "clothes_images",
}
VIDEO_LIST_FIELDS = {"videos", "video_urls", "reference_videos", "ref_videos"}
AUDIO_LIST_FIELDS = {"audios", "audio_urls", "reference_audios"}
VIDEO_FIELDS = {"video", "source_video", "target_video", "video_url", "input_video"}
AUDIO_FIELDS = {"audio", "source_audio", "target_audio", "audio_url", "music", "voice_audio"}
VOICE_TEXT_FIELDS = {"voice", "voice_id", "voice_name", "voice_description"}
BOOLEAN_NAMES = {
    "enable_base64_output",
    "enable_sync_mode",
    "prompt_enhancer",
    "enable_prompt_expansion",
    "face_enhance",
    "bgm",
}

CATEGORY_BY_CAPABILITY_PREFIX = {
    "image": "image",
    "video": "video",
    "audio": "audio",
    "speech": "audio",
    "text_to_speech": "audio",
    "text_to_audio": "audio",
    "voice": "audio",
    "talking": "avatar",
    "lip": "avatar",
    "portrait": "avatar",
    "face": "avatar",
    "text_to_3d": "3d",
    "image_to_3d": "3d",
    "llm": "llm",
    "text_generation": "llm",
    "moderation": "moderation",
    "training": "training",
}


def import_catalog(workbook_path: Path, output_path: Path, exclusions_path: Path | None = None) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    output_path = Path(output_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing = REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(sorted(missing))}")

    models_full = read_sheet(workbook["Models_Full"])
    api_schemas = {str(row.get("model_id")): row for row in read_sheet(workbook["API_Schemas"]) if row.get("model_id")}
    schema_rows = read_sheet(workbook["Schema_Fields"])
    capability_summary = read_sheet(workbook["Capability_Summary"])
    cheapest_rows = read_sheet(workbook["Cheapest_By_Capability"])
    exclusions = load_exclusions(exclusions_path)

    fields_by_model: dict[str, list[dict[str, Any]]] = {}
    for row in schema_rows:
        model_id = str(row.get("model_id") or "").strip()
        if not model_id:
            continue
        fields_by_model.setdefault(model_id, []).append(row)

    normalized: list[WaveSpeedCatalogModel] = []
    for model_row in models_full:
        model_id = str(model_row.get("model_id") or "").strip()
        if not model_id:
            continue
        schema_row = api_schemas.get(model_id, {})
        required_fields = parse_csv(schema_row.get("required_fields") or model_row.get("required_fields"))
        fields = [normalize_field(row, required_fields) for row in fields_by_model.get(model_id, [])]
        capability_tags = parse_csv(model_row.get("capability_tags"))
        primary_capability = str(model_row.get("primary_capability") or schema_row.get("primary_capability") or "other").strip()
        raw_type = str(model_row.get("raw_type") or schema_row.get("raw_type") or "").strip() or None
        category = infer_category(primary_capability, capability_tags, raw_type)
        output_kind = infer_output_kind(category, primary_capability, raw_type)
        exclusion = exclusions.get(model_id, {})
        excluded = bool(exclusion.get("excluded", False))
        record = WaveSpeedCatalogModel(
            model_id=model_id,
            display_name=str(model_row.get("name") or model_id),
            provider=clean_str(model_row.get("provider")),
            family=clean_str(model_row.get("family")),
            slug_leaf=clean_str(model_row.get("slug_leaf")),
            raw_type=raw_type,
            primary_capability=primary_capability,
            capability_tags=capability_tags or [primary_capability],
            category=category,
            output_kind=output_kind,
            base_price=to_float(model_row.get("base_price")),
            pricing_basis_guess=clean_str(model_row.get("pricing_basis_guess")),
            pricing_formula_raw=clean_str(model_row.get("pricing_formula_raw")),
            pricing_text_from_description=clean_str(model_row.get("pricing_text_from_description")),
            api_path=clean_str(schema_row.get("api_path") or model_row.get("api_path")),
            method=clean_str(schema_row.get("method") or model_row.get("method")) or "POST",
            server=clean_str(schema_row.get("server") or model_row.get("server")),
            schema_type=clean_str(schema_row.get("schema_type") or model_row.get("schema_type")),
            required_fields=required_fields,
            fields=fields,
            supports_prompt=to_bool(model_row.get("supports_prompt")),
            supports_negative_prompt=to_bool(model_row.get("supports_negative_prompt")),
            supports_image_input=to_bool(model_row.get("supports_image_input")),
            supports_video_input=to_bool(model_row.get("supports_video_input")),
            supports_audio_input=to_bool(model_row.get("supports_audio_input")),
            supports_seed=to_bool(model_row.get("supports_seed")),
            supports_prompt_expansion=to_bool(model_row.get("supports_prompt_expansion")),
            supports_base64_output=to_bool(model_row.get("supports_base64_output")),
            sort_order=int(to_float(model_row.get("sort_order")) or 1000),
            docs_url=docs_url(model_id),
            enabled=not excluded,
            enabled_reason="Loaded from WaveSpeed catalog workbook" if not excluded else "Excluded from generic runtime",
            excluded=excluded,
            exclusion_reason=str(exclusion.get("reason") or ""),
            raw_schema={
                "models_full": compact_dict(model_row),
                "api_schema": compact_dict(schema_row),
            },
        )
        normalized.append(record)

    normalized.sort(key=lambda item: (item.sort_order, item.category, item.model_id))
    payload = {
        "schema_name": "wavespeed_catalog_normalized",
        "version": 1,
        "source_workbook": str(workbook_path.as_posix()),
        "counts": {
            "models": len(normalized),
            "schema_fields": len(schema_rows),
            "capabilities": len(capability_summary),
            "cheapest_by_capability": len(cheapest_rows),
        },
        "capabilities": capability_summary,
        "cheapest_by_capability": cheapest_rows,
        "models": [model.model_dump(mode="json") for model in normalized],
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    if exclusions_path and not Path(exclusions_path).exists():
        Path(exclusions_path).parent.mkdir(parents=True, exist_ok=True)
        Path(exclusions_path).write_text("[]\n", encoding="utf-8")
    return payload["counts"]


def read_sheet(sheet) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records: list[dict[str, Any]] = []
    for values in rows:
        record = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def normalize_field(row: dict[str, Any], required_fields: list[str]) -> WaveSpeedCatalogField:
    name = str(row.get("field_name") or "").strip()
    enum_options = parse_enum_options(row.get("enum_options"))
    raw_type = clean_str(row.get("field_type"))
    field_type, asset_kind = infer_field_type(
        name=name,
        raw_type=raw_type,
        enum_options=enum_options,
        ui_component=clean_str(row.get("ui_component")),
        accept=clean_str(row.get("accept")),
    )
    return WaveSpeedCatalogField(
        name=name,
        type=field_type,
        raw_type=raw_type,
        required=to_bool(row.get("required")) or name in required_fields,
        default=parse_default(row.get("default"), field_type),
        options=enum_options,
        asset_kind=asset_kind,
        accept=clean_str(row.get("accept")),
        min_value=to_float(row.get("minimum")),
        max_value=to_float(row.get("maximum")),
        min_items=to_int(row.get("min_items")),
        max_items=to_int(row.get("max_items")),
        description=clean_str(row.get("description")) or "",
        disabled=to_bool(row.get("disabled")),
        raw_schema=compact_dict(row),
    )


def infer_field_type(
    *,
    name: str,
    raw_type: str | None,
    enum_options: list[Any],
    ui_component: str | None,
    accept: str | None,
) -> tuple[str, AssetKind | None]:
    lower = name.lower()
    raw = (raw_type or "").lower()
    ui = (ui_component or "").lower()
    accept_value = (accept or "").lower()
    if enum_options:
        return "select", None
    if ui in {"uploader", "upload"} or accept_value:
        if "image" in accept_value:
            return "asset_url", AssetKind.image
        if "video" in accept_value:
            return "asset_url", AssetKind.video
        if "audio" in accept_value:
            return "asset_url", AssetKind.audio
        return "file_url", AssetKind.other
    if lower in IMAGE_LIST_FIELDS:
        return "asset_url_list", AssetKind.image
    if lower in VIDEO_LIST_FIELDS:
        return "asset_url_list", AssetKind.video
    if lower in AUDIO_LIST_FIELDS:
        return "asset_url_list", AssetKind.audio
    if lower.endswith("_images"):
        return "asset_url_list", AssetKind.image
    if lower.endswith("_videos"):
        return "asset_url_list", AssetKind.video
    if lower.endswith("_audios"):
        return "asset_url_list", AssetKind.audio
    if lower in IMAGE_FIELDS:
        return "asset_url", AssetKind.image
    if lower in VIDEO_FIELDS:
        return "asset_url", AssetKind.video
    if lower in AUDIO_FIELDS and lower not in VOICE_TEXT_FIELDS:
        return "asset_url", AssetKind.audio
    if lower in BOOLEAN_NAMES or raw == "boolean":
        return "boolean", None
    if raw in {"integer", "int"}:
        return "integer", None
    if raw in {"number", "float", "double"}:
        return "number", None
    if raw in {"array", "list"}:
        return "json", None
    if raw in {"object", "json"}:
        return "json", None
    if lower in TEXTAREA_NAMES:
        return "textarea", None
    if raw == "string" or not raw:
        return "string", None
    return "unknown", None


def infer_category(primary_capability: str, capability_tags: list[str], raw_type: str | None) -> str:
    values = [primary_capability, *(capability_tags or []), raw_type or ""]
    joined = " ".join(value.lower().replace("-", "_") for value in values)
    if "3d" in joined:
        return "3d"
    for prefix, category in CATEGORY_BY_CAPABILITY_PREFIX.items():
        if prefix in joined:
            return category
    if "text_to_text" in joined or "chat" in joined:
        return "llm"
    return "other"


def infer_output_kind(category: str, primary_capability: str, raw_type: str | None) -> AssetKind:
    value = f"{primary_capability} {raw_type or ''}".lower().replace("-", "_")
    if "to_image" in value or category == "image":
        return AssetKind.image
    if "to_video" in value or category in {"video", "avatar"}:
        return AssetKind.video
    if "to_audio" in value or "speech" in value or category == "audio":
        return AssetKind.audio if "speech_to_text" not in value and "audio_to_text" not in value else AssetKind.other
    return AssetKind.other


def parse_csv(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).split(",") if item.strip()]


def parse_enum_options(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    text = str(value).strip()
    match = re.search(r"\[(.*)\]", text)
    if match:
        body = f"[{match.group(1)}]"
        try:
            return json.loads(body)
        except json.JSONDecodeError:
            return [item.strip().strip("\"'") for item in match.group(1).split(",") if item.strip()]
    return parse_csv(text)


def parse_default(value: Any, field_type: str) -> Any:
    if value in (None, ""):
        return None
    if field_type == "boolean":
        return to_bool(value)
    if field_type == "integer":
        return to_int(value)
    if field_type == "number":
        return to_float(value)
    return value


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> int | None:
    number = to_float(value)
    return int(number) if number is not None else None


def clean_str(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value)


def compact_dict(data: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in data.items() if value not in (None, "")}


def docs_url(model_id: str) -> str:
    if model_id.startswith(("openai/", "deepseek/")):
        return f"https://wavespeed.ai/llm/{model_id}"
    return f"https://wavespeed.ai/models/{model_id}"


def load_exclusions(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not Path(path).exists():
        return {}
    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        records = raw.get("models", [])
    else:
        records = raw
    return {str(item.get("model_id")): item for item in records if isinstance(item, dict) and item.get("model_id")}
